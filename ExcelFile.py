import openpyxl
from RefineData import RefineData
from pathlib import Path
from math import ceil

class ExcleFile():
    def __init__(self, FilePath, SheetName) -> None:
        self.FilePath = FilePath
        self.SheetName = SheetName

    def path(self,FilePath):
        relative = Path(f"Files/{FilePath}")
        absolute = relative.absolute() 
        return absolute

    # Load the Excle File
    # Parameters are FileName/FilePath and Sheet Name to scan,One Sheet at a time
    def OpenWorkBook(self, FilePath, SheetName) -> list:
        path = self.path(FilePath)
        wb = openpyxl.load_workbook(path)
        activeSheet = wb[SheetName]
        return wb, activeSheet

    # Read all the records of the file and returns in List
    def ReadAllRecords(self, usedByUser=0) -> list:

        FullData = []  # Defining an Empty list, which later will appends all the records in a list element and used as a return value
        i = 2  # If we put it 1 then it will also count heading cell in the print statement for the Record Completed
        WorkBook = self.OpenWorkBook(
            self.FilePath, self.SheetName)  # Opening the Excel File
        activeSheet = WorkBook[1]  # Only using defined of the Workbook
        # Total rows in a Workbook (+1 because activeSheet.max_row starts with 0)
        rows = activeSheet.max_row+1
        # Total column in a Workbook (+1 because activeSheet.max_column starts with 0)
        column = activeSheet.max_column+1

        # Using Nested For loop to cover the full table
        for y in range(2, rows):  # y-axis
            data_lst = []  # This variable will store the individual record i.e.['VF456VGFDBD4564654',465897489,-,-,-,NONE...]
            for x in range(1, column):  # x-axis
                cell_obj = activeSheet.cell(row=y, column=x)  # Extract the raw value of the cell in a variable
                data_lst.append(cell_obj.value) # Appending to the desired value in data_lst list
            
            # Appending the datalist as [['gdsg',46546,...],['gdsg',46546,...],['gdsg',46546,...]...]
            FullData.append(data_lst)
            if usedByUser == 0:
                print(f"Record completed {ceil((i/rows)*100)}%")
            i += 1
        print(f"Stored Data Length : {len(FullData)}")
        return FullData

    def refine_data(self,data_lst,num,errorLst):
        rd = RefineData(data_lst,num,errorLst)
        return rd.filter()[0]

    def FilterAndInsertRecords(self, NewFileName="Book1.xlsx", SheetName="Sheet1"):
        errorLst = []
        unFiltered_data, num = self.ReadAllRecords(usedByUser=1), 2
        newWorkBook = self.OpenWorkBook(NewFileName, SheetName)
        ActiveSheet = newWorkBook[1]
        for data in unFiltered_data:
            try:
                ActiveSheet.append(self.refine_data(data,num,errorLst))
            except Exception as e:
                errorLst.append(f"Line No. {num} Error : {e}")
            num += 1
        newWorkBook[0].save((NewFileName))
        self.Error_Output(errorLst)
        print(f"Completed records",num)

    def Error_Output(self,error_lst):
        file = open("Error.txt","w")
        for error in error_lst:
            file.writelines(error+"\n")
        file.close()

if __name__ == '__main__':
    ExFile = ExcleFile("test_original.xlsx", "Build")
    ExFile.FilterAndInsertRecords("Book2.xlsx")